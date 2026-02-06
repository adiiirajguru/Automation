from openpyxl import Workbook, load_workbook
from pathlib import Path


def main():
    base_dir = Path(__file__).resolve().parent
    excel_file = base_dir / "demo.xlsx"

    # ---------- CREATE EXCEL ----------
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Employees"

    # Write header
    sheet.append(["Name", "Age", "City"])

    # Write rows
    sheet.append(["Aditya", 23, "Indore"])
    sheet.append(["Rahul", 21, "Pune"])
    sheet.append(["Priya", 25, "Bangalore"])

    wb.save(excel_file)

    print("Excel file created!")

    # ---------- READ + MODIFY ----------
    wb = load_workbook(excel_file)
    sheet = wb["Employees"]

    print("\nReading Excel data:")

    for row in sheet.iter_rows(values_only=True):
        print(row)

    # Modify a cell
    sheet["B2"] = 24  # change age

    wb.save(excel_file)

    print("\n Excel updated!")


if __name__ == "__main__":
    main()
