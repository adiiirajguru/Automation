from openpyxl import load_workbook
from pathlib import Path


def main():
    base_dir = Path(__file__).resolve().parent
    excel_file = base_dir / "employees.xlsx"

    wb = load_workbook(excel_file)
    sheet = wb["Employees"]

    print("Excel_CSV_file_cleaner contents:")

    for row in sheet.iter_rows(values_only=True):
        print(row)


if __name__ == "__main__":
    main()
